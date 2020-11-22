import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
excel_document = openpyxl.load_workbook('DevExercise.xlsx')
    
def SaveFile():
    excel_document.save('DevExercise1.xlsx')
    
def ColorMax(all_columns,all_rows,sheet):
        #finding max
    blue = PatternFill(start_color='ff296d98', end_color='ff296d98',fill_type='solid')
    max = -1
    max_flag = -1
    
    for column in range(len(all_columns)):
        for row in range(len(all_rows)):
            if column == len(all_columns)-1 and row > 0:
                if sheet.cell(row = (row+1), column = (column+2)).value > max:
                    max = sheet.cell(row = (row+1), column = (column+2)).value
                    max_flag = row+1
    
    for column in range(len(all_columns)):
        sheet.cell(row = (max_flag), column = (column+1)).fill = blue

def CalcSum(sheets):
    sheet = excel_document[sheets]
    columns = sheet.columns
    all_columns = []
    rows = sheet.rows
    all_rows = []

    for col in columns:
        all_columns.append(col)
        
    for row in rows:
        all_rows.append(row)

    #Calculating sum
    sum = 0
    #set yellow color
    yellow = PatternFill(start_color='ffffff00', end_color='ffffff00',fill_type='solid')
    rent_flag = -1
    
    for row in range(len(all_rows)):
        for col in range(len(all_columns)):
            if all_rows[row][col].value == 'Rent':
                all_rows[row][col-1].fill = yellow
                rent_flag = row
            if row == rent_flag:
                all_rows[row][col].fill = yellow
            if col > 1 and row > 0:
                if type(all_rows[row][col].value) == int:
                    sum += all_rows[row][col].value
            if col == len(all_columns)-1 and row > 0:
                sheet.cell(row = (row+1), column = (col+2), value = sum)
                sum = 0
                
    ColorMax(all_columns,all_rows,sheet)

    return rent_flag
    
def AddRentTotalTab(rentFlags ,index, sheets ):
    sheet = excel_document[sheets]
    columns = sheet.columns
    all_columns = []
    rows = sheet.rows
    all_rows = []

    for col in columns:
        all_columns.append(col)
        
    for row in rows:
        all_rows.append(row)
        
    all_columns_len = len(all_columns)-1
    newSheet = excel_document["RentTotal"]
    
    newSheet.cell(row = index+1, column = 1, value = sheets)
    newSheet.cell(row = index+1, column = 2 , value = all_rows[rentFlags[index]][all_columns_len].value)

def main():
    # getting sheets.
    sheets_arr = []
    rentFlags = []
    for sheet in excel_document.sheetnames:
        sheets_arr.append(sheet)
        
    # Calling to the CalcSum function and save the result at rentFlags array
    for sheets in sheets_arr:   
        rentFlags.append(CalcSum(sheets))
        
    index = 0
    excel_document.create_sheet("RentTotal")
    for sheets in sheets_arr:
        AddRentTotalTab(rentFlags, index ,sheets)
        index+=1
        
    SaveFile()

if __name__ == "__main__":
    main()